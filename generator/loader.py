from dataclasses import dataclass
from pathlib import Path
import csv

from openpyxl import load_workbook


@dataclass
class Region:
    province: str
    city: str
    district: str
    dong: str


class DataLoader:
    REQUIRED_COLUMNS = {
        "시도": "province",
        "시군구": "city",
        "구": "district",
        "동": "dong",
    }

    def load(self, filepath):
        if not filepath:
            return []

        filepath = Path(filepath).resolve()
        if filepath.is_dir():
            return self.load_directory(filepath)

        workbook = load_workbook(filepath, read_only=True, data_only=True)
        worksheet = workbook.active

        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        has_header = self._has_header(first_row)
        column_indexes = self._column_indexes(first_row) if has_header else {
            column: index for index, column in enumerate(self.REQUIRED_COLUMNS)
        }
        min_row = 2 if has_header else 1

        regions = []
        for row in worksheet.iter_rows(min_row=min_row, values_only=True):
            values = {
                field: self._cell_value(self._row_value(row, index))
                for column, field in self.REQUIRED_COLUMNS.items()
                for index in [column_indexes[column]]
            }

            if any(values.values()):
                regions.append(
                    Region(
                        province=values["province"],
                        city=values["city"],
                        district=values["district"],
                        dong=values["dong"],
                    )
                )

        workbook.close()
        return regions

    def load_directory(self, directory):
        directory = Path(directory).resolve()
        if not directory.exists():
            return []

        regions = []
        for path in sorted(directory.rglob("*")):
            if not path.is_file() or path.name.startswith("~$"):
                continue

            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                regions.extend(self.load(path))
            elif path.suffix.lower() == ".csv":
                regions.extend(self._load_csv(path))

        return self._unique_regions(regions)

    def _load_csv(self, filepath):
        regions = []
        with Path(filepath).open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            if not reader.fieldnames:
                return regions

            if not all(column in reader.fieldnames for column in self.REQUIRED_COLUMNS):
                return regions

            for row in reader:
                values = {
                    field: self._cell_value(row.get(column))
                    for column, field in self.REQUIRED_COLUMNS.items()
                }
                if any(values.values()):
                    regions.append(Region(**values))

        return regions

    def _unique_regions(self, regions):
        seen = set()
        unique = []
        for region in regions:
            key = (region.province, region.city, region.district, region.dong)
            if key in seen:
                continue
            seen.add(key)
            unique.append(region)
        return unique

    def _column_indexes(self, header):
        columns = {name: index for index, name in enumerate(header)}
        missing_columns = [
            column for column in self.REQUIRED_COLUMNS if column not in columns
        ]

        if missing_columns:
            names = ", ".join(missing_columns)
            raise ValueError(f"Missing required columns: {names}")

        return columns

    def _has_header(self, row):
        values = set(row)
        return all(column in values for column in self.REQUIRED_COLUMNS)

    def _row_value(self, row, index):
        if index >= len(row):
            return None

        return row[index]

    def _cell_value(self, value):
        if value is None:
            return ""

        return str(value)

from collections import Counter
from dataclasses import dataclass
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook

from config import OUTPUT_DIR, PROJECT_ROOT


ERROR_VALUES = {"#ERROR!", "NULL", "NONE", "NAN", "#####"}
MIN_CONTENT_LENGTH = 2500
MIN_GENERATED_LENGTH = 2800
MAX_GENERATED_LENGTH = 3500
CONTENT_SOURCE = "새지역만들기.xlsx"


@dataclass
class ContentResult:
    page: object
    content_key: str
    status: str
    reasons: list
    source: str
    length: int
    generated_length: int = 0
    error: str = ""


class ContentFallbackGenerator:
    def __init__(
        self,
        content_filepath=CONTENT_SOURCE,
        report_path=None,
    ):
        self.content_filepath = Path(content_filepath)
        if not self.content_filepath.is_absolute():
            self.content_filepath = PROJECT_ROOT / self.content_filepath
        self.content_filepath = self.content_filepath.resolve()
        self.report_path = Path(report_path or OUTPUT_DIR).resolve() / "content_generation_report.xlsx"
        self.results = []
        self.duplicate_content_keys = []

    def build(self, pages):
        page_list = self._flatten(pages)
        content_map = self._load_content()
        results = []

        for page in page_list:
            content_key = self._content_key(page)
            source_key, source_content = self._lookup_content(content_map, page, content_key)
            source_content = self._normalize_fragment(source_content)
            reasons = self._content_reasons(source_content)

            if not reasons:
                page.content = source_content
                results.append(
                    ContentResult(
                        page=page,
                        content_key=source_key,
                        status="EXCEL",
                        reasons=[],
                        source="excel",
                        length=self._content_length(source_content),
                    )
                )
                continue

            generated = self._normalize_fragment(self._generate(page, source_content))
            validation_errors = self._validate_generated(page, source_content, generated)

            if validation_errors:
                page.content = source_content
                results.append(
                    ContentResult(
                        page=page,
                        content_key=source_key,
                        status="ERROR",
                        reasons=reasons,
                        source="excel",
                        length=self._content_length(source_content),
                        generated_length=self._content_length(generated),
                        error=", ".join(validation_errors),
                    )
                )
                continue

            page.content = generated
            results.append(
                ContentResult(
                    page=page,
                    content_key=source_key,
                    status="GENERATED",
                    reasons=reasons,
                    source="fallback",
                    length=self._content_length(source_content),
                    generated_length=self._content_length(generated),
                )
            )

        self.results = results
        self._write_report(results)
        return results

    def _lookup_content(self, content_map, page, content_key):
        if content_key in content_map:
            return content_key, content_map[content_key]

        if page.title in content_map:
            return page.title, content_map[page.title]

        return content_key, ""

    def _load_content(self):
        if not self.content_filepath.exists():
            return {}

        workbook = load_workbook(self.content_filepath, read_only=True, data_only=True)
        content = {}
        duplicates = []

        for worksheet in workbook.worksheets:
            first_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )
            column_map = self._column_map(first_row)
            min_row = 2 if column_map else 1

            for row in worksheet.iter_rows(min_row=min_row, values_only=True):
                if not row or not row[0]:
                    continue

                content_key, body, enforce_duplicate = self._row_content(row, column_map)

                if not content_key:
                    continue

                if content_key in content and enforce_duplicate:
                    duplicates.append(content_key)
                    continue

                if content_key not in content:
                    content[content_key] = body

        workbook.close()

        if duplicates:
            self.duplicate_content_keys = sorted(set(duplicates))
            names = "\n".join(self.duplicate_content_keys)
            raise ValueError(f"Duplicate Content Key\n{names}")

        self.duplicate_content_keys = []
        return content

    def _column_map(self, row):
        if not row:
            return None

        columns = {str(value).strip(): index for index, value in enumerate(row) if value}
        if "Content Key" in columns and "Content" in columns:
            return {
                "key": columns["Content Key"],
                "content": columns["Content"],
                "enforce_duplicate": True,
            }

        if "제목" in columns and "Content" in columns:
            return {
                "key": columns["제목"],
                "content": columns["Content"],
                "enforce_duplicate": False,
            }

        return None

    def _row_content(self, row, column_map):
        if column_map:
            content_key = self._row_value(row, column_map["key"])
            body = self._row_value(row, column_map["content"])
            return content_key, body, column_map["enforce_duplicate"]

        content_key = self._row_value(row, 0)
        body = self._row_value(row, 1)
        return content_key, body, "/" in content_key

    def _row_value(self, row, index):
        if index >= len(row):
            return ""

        value = row[index]
        if value is None:
            return ""

        return str(value).strip()

    def _normalize_fragment(self, content):
        text = self._text(content)
        text = re.sub(r"(?is)<!doctype[^>]*>", "", text)
        text = re.sub(r"(?is)<head\b[^>]*>.*?</head>", "", text)
        text = re.sub(r"(?is)</?html\b[^>]*>", "", text)
        text = re.sub(r"(?is)</?body\b[^>]*>", "", text)
        text = re.sub(r"(?m)(^|\n)(\s*)h([1-6])>", r"\1\2<h\3>", text)
        return text.strip()

    def _content_key(self, page):
        nodes = []
        current = page

        while current is not None:
            if getattr(current, "title", None):
                nodes.append(current.title)
            current = getattr(current, "parent", None)

        nodes.reverse()
        return "/".join(nodes)

    def _flatten(self, pages):
        roots = pages if isinstance(pages, list) else [pages]
        flattened = []

        for page in roots:
            self._walk(page, flattened)

        return flattened

    def _walk(self, page, flattened):
        if page is None:
            return

        flattened.append(page)

        for child in getattr(page, "children", []):
            self._walk(child, flattened)

    def _content_reasons(self, content):
        reasons = []
        stripped = self._text(content).strip()
        upper = stripped.upper()

        if not stripped:
            reasons.append("EMPTY")
            return reasons

        if upper in ERROR_VALUES:
            reasons.append("ERROR")
            return reasons

        if self._has_broken_html(stripped):
            reasons.append("BROKEN_HTML")

        if self._content_length(stripped) < MIN_CONTENT_LENGTH:
            reasons.append("SHORT_CONTENT")

        if not self._renders_content(stripped):
            reasons.append("RENDER_ERROR")

        return reasons

    def _generate(self, page, source_content):
        region = self._region_name(page)
        subject = self._subject_name(page)
        keyword_count = self._keyword_count(source_content, page.title)
        title_phrase = self._title_phrase(page.title, keyword_count)

        sections = [
            (
                "학습 상황 점검",
                [
                    f"{region} 지역의 {subject} 학습은 학생이 이미 이해한 부분과 아직 흔들리는 부분을 구분하는 데서 시작합니다. 학교 수업에서 다룬 개념, 최근 과제에서 반복된 실수, 시험 범위 안에서 부담을 느끼는 단원을 함께 확인하면 필요한 학습 순서를 정하기 쉽습니다.",
                    f"수업 전에는 지난 학습 내용을 짧게 확인하고, 수업 중에는 개념과 문제 풀이가 어떻게 연결되는지 살피는 과정이 필요합니다. 학생이 답을 맞혔더라도 설명이 불안정하면 다음 단계에서 같은 유형을 다시 틀릴 수 있으므로 풀이 과정을 말로 정리하게 하는 것이 좋습니다.",
                ],
            ),
            (
                "수업 구성 기준",
                [
                    f"{subject} 수업은 개념 정리, 대표 유형 풀이, 오답 분석, 유사 문제 확인, 다음 학습 범위 안내의 흐름으로 구성하면 안정적입니다. 이 순서가 유지되면 학생은 무엇을 배웠고 무엇을 다시 봐야 하는지 분명하게 알 수 있습니다.",
                    f"기초가 부족한 학생은 쉬운 내용부터 다시 점검하는 과정이 필요하고, 이미 이해가 빠른 학생은 풀이 속도와 표현 방식을 다듬는 과정이 중요합니다. 같은 지역의 학생이라도 학년, 학교 일정, 평소 학습 습관에 따라 필요한 수업 방식은 달라질 수 있습니다.",
                ],
            ),
            (
                "오답 관리",
                [
                    "오답은 단순히 틀린 문제를 다시 푸는 자료가 아니라 학생의 학습 상태를 보여 주는 기록입니다. 개념 부족, 조건 누락, 계산 실수, 시간 관리, 서술 방식의 문제처럼 원인을 나누어 보면 다음 수업에서 무엇을 먼저 다룰지 판단하기 쉽습니다.",
                    "같은 실수가 반복되는 경우에는 문제 수를 늘리기보다 풀이 습관을 먼저 확인해야 합니다. 문제를 읽는 순서, 조건을 표시하는 방식, 답을 검산하는 시간, 해설을 읽은 뒤 다시 설명하는 과정이 갖춰지면 학습의 안정성이 높아집니다.",
                ],
            ),
            (
                "학습 계획",
                [
                    f"{region}의 학습 계획은 학생의 생활 리듬과 학교 일정을 함께 고려해야 합니다. 수행평가와 시험 기간이 겹치는 시기에는 과제량을 조절하고, 평소에는 복습 주기를 짧게 유지해 배운 내용을 오래 남기는 편이 좋습니다.",
                    "무리하게 많은 내용을 한 번에 다루면 학생은 무엇을 우선해야 하는지 혼란을 느낄 수 있습니다. 이번 수업에서 확인할 단원, 이번 주에 줄일 실수, 다음 시험 전까지 완성할 범위를 작게 나누면 학습 목표가 더 현실적입니다.",
                ],
            ),
            (
                "가정 학습 연결",
                [
                    "수업에서 다룬 내용이 가정 학습으로 이어지려면 기록이 간단하고 분명해야 합니다. 배운 개념, 다시 풀 문제, 자주 틀린 이유, 다음에 확인할 단원을 나누어 정리하면 학생도 혼자 공부할 때 기준을 잃지 않습니다.",
                    "보호자가 학습 상태를 확인할 때도 결과만 보는 것보다 과정의 변화를 살피는 편이 도움이 됩니다. 어떤 부분을 정확히 이해했고, 어떤 부분은 아직 보완이 필요한지 알 수 있으면 다음 학습 방향을 더 차분하게 잡을 수 있습니다.",
                ],
            ),
        ]

        html_parts = ["<section>", f"<h2>{title_phrase}학습 안내</h2>", "<article>"]

        for heading, paragraphs in sections:
            html_parts.append(f"<h3>{heading}</h3>")
            for paragraph in paragraphs:
                html_parts.append(f"<p>{paragraph}</p>")

        html_parts.extend(
            [
                "<h3>확인할 항목</h3>",
                "<ul>",
                "<li>학교 진도와 학생의 현재 이해도를 함께 확인합니다.</li>",
                "<li>오답 원인을 분류해 다음 수업의 우선순위를 정합니다.</li>",
                "<li>복습 범위와 과제량을 학생의 생활 리듬에 맞게 조절합니다.</li>",
                "<li>수업 후 스스로 다시 설명할 수 있는지 점검합니다.</li>",
                "</ul>",
                "</article>",
                "</section>",
            ]
        )

        return self._fit_length("\n".join(html_parts), page, keyword_count)

    def _fit_length(self, content, page, keyword_count):
        additions = [
            "학습 변화는 한 번의 설명보다 꾸준한 확인에서 더 분명하게 나타납니다. 학생이 스스로 이해한 내용을 말로 설명하고, 다시 풀어야 할 문제를 구분하며, 다음 수업에서 확인할 기준을 알게 되면 수업의 흐름이 안정됩니다.",
            "수업 자료는 학생이 실제로 다시 볼 수 있는 형태여야 합니다. 너무 많은 내용을 한꺼번에 적기보다 핵심 개념, 대표 실수, 다시 풀 문제, 다음 학습 범위를 구분하면 복습 부담을 줄일 수 있습니다.",
            "시험을 앞둔 시기에는 새로운 내용을 무리하게 늘리기보다 이미 배운 범위에서 흔들리는 부분을 정리하는 것이 좋습니다. 자주 틀리는 유형을 확인하고 실전에서 시간을 어떻게 배분할지 함께 살피면 준비 과정이 더 분명해집니다.",
            "학생마다 자신감을 잃는 지점은 다릅니다. 어떤 학생은 문제를 읽는 단계에서 막히고, 어떤 학생은 개념은 알지만 답안 작성이 불안정하며, 또 다른 학생은 복습 간격이 길어져 배운 내용을 금방 잊기도 합니다.",
            "수업의 난이도는 학생이 견딜 수 있는 속도 안에서 조절되어야 합니다. 너무 쉬우면 집중이 떨어지고, 너무 어려우면 시도 자체를 피하게 되므로 현재 수준보다 조금 높은 과제를 통해 사고 과정을 넓히는 방식이 필요합니다.",
            "복습은 오래 붙잡는 것보다 자주 확인하는 편이 효과적입니다. 짧은 시간이라도 지난 수업의 핵심을 다시 떠올리고, 같은 유형을 한두 문제 풀어 보며, 설명이 막히는 부분을 표시하면 다음 수업의 출발점이 분명해집니다.",
            "학습 기록은 학생과 보호자가 같은 기준으로 상태를 이해하게 돕습니다. 단순히 숙제를 했는지보다 어떤 실수가 줄었고 어떤 부분이 남아 있는지 확인하면 수업의 방향을 차분하게 조정할 수 있습니다.",
            "학교 수업과 별개의 내용을 무리하게 진행하면 학생이 체감하는 부담이 커질 수 있습니다. 현재 학교 진도와 필요한 보완 단원을 함께 놓고 보면서 우선순위를 정해야 학습이 끊기지 않고 이어집니다.",
            "문제를 맞힌 뒤에도 풀이가 안정적인지 다시 살피는 과정이 필요합니다. 우연히 답을 찾은 문제와 개념을 이해해 해결한 문제는 다음 단원에서 차이가 드러나므로, 답보다 풀이 근거를 확인하는 습관이 중요합니다.",
            "학생이 질문을 어려워하는 경우에는 틀린 문제를 바로 지적하기보다 어느 단계에서 막혔는지 함께 확인하는 방식이 좋습니다. 질문이 자연스러워지면 수업 중 피드백도 빨라지고, 부족한 부분을 숨기지 않게 됩니다.",
            "장기적인 학습에서는 작은 성공 경험을 꾸준히 쌓는 것이 중요합니다. 한 단원의 핵심을 이해하고, 비슷한 문제를 스스로 해결하며, 이전보다 실수가 줄어드는 경험이 반복되면 다음 학습으로 넘어갈 힘이 생깁니다.",
            "본문의 목적은 특정 결과를 약속하는 것이 아니라 학습을 판단하는 기준을 분명히 보여 주는 데 있습니다. 학생의 현재 위치를 확인하고 필요한 보완 과정을 정리하면 수업 내용도 더 현실적인 방향으로 구성됩니다.",
        ]

        index = 0
        while self._content_length(content) < MIN_GENERATED_LENGTH:
            if index < len(additions):
                paragraph = additions[index]
            else:
                point = index - len(additions) + 1
                paragraph = (
                    f"추가 점검 {point}에서는 수업 전후의 변화를 짧게 기록합니다. "
                    f"학생이 어떤 문제에서 멈췄는지, 어떤 설명 뒤에 이해가 나아졌는지, "
                    f"다음 시간에 다시 확인할 부분이 무엇인지 남겨 두면 학습 흐름을 끊지 않고 이어 갈 수 있습니다."
                )
            insert = f"<p>{paragraph}</p>"
            content = content.replace("</article>", f"{insert}\n</article>", 1)
            index += 1
            if index > 30:
                break

        if self._content_length(content) > MAX_GENERATED_LENGTH:
            return content

        return content

    def _validate_generated(self, page, source_content, generated):
        errors = []
        length = self._content_length(generated)

        if length < MIN_GENERATED_LENGTH:
            errors.append("TOO_SHORT")

        if length > MAX_GENERATED_LENGTH:
            errors.append("TOO_LONG")

        if self._has_broken_html(generated):
            errors.append("BROKEN_HTML")

        if self._keyword_count(generated, page.title) != self._keyword_count(source_content, page.title):
            errors.append("KEYWORD_COUNT_CHANGED")

        region = self._region_name(page)
        if region and region not in generated:
            errors.append("REGION_MISSING")

        if self._has_duplicate_paragraph(generated):
            errors.append("DUPLICATE_PARAGRAPH")

        if self._has_sentence_cutoff(generated):
            errors.append("SENTENCE_CUTOFF")

        return errors

    def _write_report(self, results):
        self.report_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "summary"

        counts = Counter(result.status for result in results)
        summary.append(["metric", "value"])
        summary.append(["total_pages", len(results)])
        summary.append(["excel_pages", counts["EXCEL"]])
        summary.append(["generated_pages", counts["GENERATED"]])
        summary.append(["error_pages", counts["ERROR"]])

        details = workbook.create_sheet("details")
        details.append(
            [
                "title",
                "content_key",
                "url",
                "page_type",
                "status",
                "reason",
                "source_length_no_space",
                "generated_length_no_space",
                "error",
            ]
        )

        for result in results:
            details.append(
                [
                    result.page.title,
                    result.content_key,
                    result.page.url,
                    result.page.page_type.name,
                    result.status,
                    ", ".join(result.reasons),
                    result.length,
                    result.generated_length,
                    result.error,
                ]
            )

        workbook.save(self.report_path)
        workbook.close()

    def _text(self, content):
        return "" if content is None else str(content)

    def _content_length(self, content):
        plain = re.sub(r"<[^>]*>", "", unescape(self._text(content)))
        return len(re.sub(r"\s+", "", plain))

    def _renders_content(self, content):
        plain = re.sub(r"<[^>]*>", "", unescape(self._text(content))).strip()
        return bool(plain)

    def _has_broken_html(self, content):
        parser = _BalancedTagParser()
        parser.feed(self._text(content))
        parser.close()
        return parser.has_error

    def _has_duplicate_paragraph(self, content):
        paragraphs = re.findall(r"<p(?:\s[^>]*)?>.*?</p>", self._text(content), re.IGNORECASE | re.DOTALL)
        return len(paragraphs) != len(set(paragraphs))

    def _has_sentence_cutoff(self, content):
        plain = re.sub(r"<[^>]*>", "", self._text(content)).strip()
        return bool(plain) and plain[-1] not in ".!?)다요니다습니다"

    def _keyword_count(self, content, keyword):
        return self._text(content).count(keyword)

    def _region_name(self, page):
        region = getattr(page, "region", None)
        for attr in ("dong", "district", "city", "province"):
            value = getattr(region, attr, "")
            if value:
                return value
        return "전국"

    def _subject_name(self, page):
        title = getattr(page, "title", "")
        for subject in ("영어", "수학", "국어", "과학"):
            if subject in title:
                return subject
        return "과외"

    def _title_phrase(self, title, count):
        if count <= 0:
            return ""

        return f"{title} "


class _BalancedTagParser(HTMLParser):
    CHECK_TAGS = {"section", "article", "h2", "h3", "p", "ul", "li", "strong", "em"}

    def __init__(self):
        super().__init__()
        self.stack = []
        self.has_error = False

    def handle_starttag(self, tag, attrs):
        if tag in self.CHECK_TAGS:
            self.stack.append(tag)

    def handle_endtag(self, tag):
        if tag not in self.CHECK_TAGS:
            return

        if not self.stack or self.stack[-1] != tag:
            self.has_error = True
            return

        self.stack.pop()

    def close(self):
        super().close()
        if self.stack:
            self.has_error = True
